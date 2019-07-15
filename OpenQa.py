from gevent import monkey

monkey.patch_all()
import os
from gevent.pool import Pool
import json, time, jsonpath, os, requests, xlwt, openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import csv
from functools import partial
import random

test_level = 1  # 1：entity + intent   2: only att   3: att + number
file_type = 1  # 1:txt  2:csv
server = 1  # 1:7.44   2:AWS


def starttestcase(file_name):
    # 请求问题的列表
    qu = []
    # 请求状态的列表
    di = []
    # entity 实体列表
    en = []
    # 期望的intent  意图
    exp_int = []
    # 期望的entity 实体
    exp_ent = []
    # intents 意图列表
    inte = []
    # value值列表
    va = []
    # 用于判断intent
    status = []

    en_statu = []

    # 用于添加解析引擎类型
    ParserEngineType = []
    threads = []

    # 用于请求时间的列表
    parser_time = []
    # 用于响应的时间列表
    answer_time = []

    # 用于添加ParserConceptType
    concept_type = []
    # 用于添加Parameter
    parameter = []

    if server == 1:
        baseurl = 'http://109.105.7.44:'
        r = 9090 + random.randint(0, 9)
        url = baseurl + str(r) + '/api/question?q='

    else:
        url = r'http://qa-zh-dev.s-knowledge.net/api/question?q='

    file = open(file_dir + file_name, encoding="utf-8")
    for eachLine in file:
        if test_level == 1:
            question = eachLine.split(',')[0].encode('utf-8').decode('utf-8-sig').strip()
            att = eachLine.split(',')[2].encode('utf-8').decode('utf-8-sig').strip()
            # exp_int.append(att)
            ent = eachLine.split(',')[1].encode('utf-8').decode("utf-8-sig").strip()
            # exp_ent.append(ent)
        elif test_level == 3:
            question = eachLine.split(',')[0].encode('utf-8').decode('utf-8-sig').strip()
            att = eachLine.split(',')[0].encode('utf-8').decode('utf-8-sig').strip()
            if file_type == 1:
                ent = file_name.split(".txt")[0]
            else:
                ent = file_name.split(".csv")[0]
        else:
            question = eachLine.split('\t')[0].encode('utf-8').decode('utf-8-sig').strip()
            att = eachLine.split('\t')[0].encode('utf-8').decode('utf-8-sig').strip()
            if file_type == 1:
                ent = file_name.split(".txt")[0]
            else:
                ent = file_name.split(".csv")[0]

        print(att)
        wurl = url + question + "&key=jamlive"
        i = 0
        print(wurl)
        v = []
        while True:
            try:
                flag = True
                if i > 0:
                    print("尝试%d次  问题：%s  连接：%s" % (i, question, wurl))
                r = requests.get(wurl)

            except:
                flag = False
                i += 1
            if flag == True:
                break

        if r.status_code >= 400:
            print("Question: %s" % question)
            print("Intent: Unknown")
            print("Status: server error")
            print("Diagnosis: fail %d" % r.status_code)
            print("Value: Empty")
            print("Entity: Empty")
            continue

        if test_level == 1:
            exp_int.append(att)
            exp_ent.append(ent)
        else:
            exp_int.append(ent)

        text = r.content.decode("utf-8")
        try:
            jsontxt = json.loads(text)
        except ValueError as e:
            print(e.message)
            print("load text error :", text)

        # 请求状态
        Diag = jsontxt["Diagnosis"]
        # 请求的时间
        par_time = jsontxt["ParsingTime_ms"]
        # 响应的时间
        ans_time = jsontxt["AnsweringTime_ms"]
        # 请求的实体
        Entity = jsontxt["Query"]["Entity"]
        if Entity == None:
            Entity = "Null"

        try:
            intents = jsontxt["Query"]["Attributes"][0]
        except:
            intents = jsontxt["Query"]["Query"]
            intent2 = intents
            if question == intent2:
                intents = "Null"
        Concept_Type = jsontxt["Query"]["ParserConceptType"]
        # 获取请求问题的内容  即Value值
        Result = jsontxt["Result"]
        if Result != None:
            AnswerType = jsontxt["Result"]["AnswerType"]
            # AnswerTypes = ["SingleKeyValue", "SingleText", "KeyValueList", "Entity", "Tabular", "EntityListAnswer"]
            # AnswerType 的第一种情况
            if AnswerType == "Entity":
                try:
                    value = jsontxt["Result"]["Items"][0]["Abstract"]
                    v.append(value)
                except:
                    value = "Null"
                    v.append(value)
            # AnswerType 的第二种情况
            elif AnswerType == "SingleText":
                if intents == "shortnoun":
                    try:
                        value = jsontxt["Result"]["Items"][-1]["Abstract"]
                        v.append(value)
                    except:
                        value = "Null"
                        v.append(value)
                else:
                    try:
                        if server == 2 and (
                                intents == "IntroductionOfTheory" or intents == "DefinitionOfThings" or intents == "WhoIsPerson"):
                            value = jsontxt["Result"]["Items"][-1]["Abstract"]
                        else:
                            value = jsontxt["Result"]["Items"][0]["Content"]
                        v.append(value)
                    except:
                        value = "Null"
                        v.append(value)
            # AnswerType 的第三种情况
            elif AnswerType == "SingleKeyValue":
                if intents == "WhoIsPerson":
                    try:
                        value = jsontxt["Result"]["Items"][-1]["Abstract"]
                        v.append(value)
                    except:
                        value = "Null"
                        v.append(value)
                else:
                    try:
                        value = jsontxt["Result"]["Items"][0]["Items"][0]["Value"]
                        v.append(value)
                    except:
                        value = "Null"
                        v.append(value)
            # AnswerType 的第四种情况
            elif AnswerType == "KeyValueList":
                try:
                    value = jsontxt["Result"]["Items"][0]["Items"][-1]["Value"]
                    v.append(value)
                except:
                    value = "Null"
                    v.append(value)
            # AnswerType 的第五种情况
            elif AnswerType == "Tabular":
                try:
                    for i in range(len(jsontxt["Result"]["Items"][0]["Items"][0]["Attributes"])):
                        key = jsontxt["Result"]["Items"][0]["Items"][0]["Attributes"][i]["Key"]
                        value = jsontxt["Result"]["Items"][0]["Items"][0]["Attributes"][i]["Value"]
                        v.append(key + value + " ")
                        print(" Tabular value = " + value)
                except:
                    print(" Tabular value = null")
                    value = "Null"
                    v.append(value)
            elif AnswerType == "EntityList":
                try:
                    for i in range(len(jsontxt["Result"]["Items"][0]["Items"])):
                        value = jsontxt["Result"]["Items"][0]["Items"][i]["Title"]
                        v.append(value)
                except:
                    value = "Null"
                    v.append(value)
                    
        try:
            ParserEn_Type = jsontxt["Query"]["ParserEngineType"]
        except:
            ParserEn_Type = "Null"
        # 获取parameter & parameterType
        para = ""
        if jsontxt["Query"]["Parameters"] != None:
            try:
                pUnit = jsontxt["Query"]["Parameters"]
                for k, w in pUnit.items():
                    para += k + ":" + w + " "
            except:
                para = None

        else:
            para = None

        value = ""
        # print(Diag)
        if len(v) > 0:
            for valu in v:
                try:
                    # value += valu + ","
                    value += valu
                except Exception as e:
                    value = "Null"
            print(value)

            qu.append(question)
            # di.append(Diag)
            en.append(Entity)
            inte.append(intents)
            va.append(value)
            ParserEngineType.append(ParserEn_Type)
            parser_time.append(par_time)
            answer_time.append(ans_time)
            concept_type.append(Concept_Type)
            parameter.append(para)


        else:
            value = "Null"
            # if values == None:
            #     value = ""
            # else:
            #     try:
            #         value = values.replace("-", "—").strip("\n ")
            #     except Exception as e:
            #         value = "Null"

            qu.append(question)
            # di.append(Diag)
            en.append(Entity)
            inte.append(intents)
            va.append(value)
            ParserEngineType.append(ParserEn_Type)
            parser_time.append(par_time)
            answer_time.append(ans_time)
            concept_type.append(Concept_Type)
            parameter.append(para)

        if test_level == 1:
            if ent == Entity:
                en_statu.append('True')
            elif Entity == 'Null':
                if att == "FamousAphorism":
                    en_statu.append('True')
                else:
                    en_statu.append('Null')
                    Diag = 'Parser fail'
            else:
                if intents != "hotfix":
                    en_statu.append('False')
                    Diag = 'Parser fail'
                else:
                    en_statu.append('Null')
        ######    判断intent    ######
        if test_level == 1:
            if intents == att:
                # status 用来判断intent  statu 用来判断value
                status.append("True")

            elif Diag == "Successful" and intents != att:
                if att == "EffectOfFood" and intents == "EffectOfNutrientOrDrug":
                    status.append("True")
                elif att == "EffectOfNutrientOrDrug" and intents == "EffectOfFood":
                    status.append("True")
                elif intents == "hotfix":
                    status.append("True")
                else:
                    status.append("False")
                    Diag = 'Parser fail'
                # statu.append('False')

                # statu.append('Null')
            elif Diag == 'Parser fail':
                status.append("Null")
                if att == "shortnoun":
                    Diag = 'shortnoun'
                else:
                    Diag = 'Parser fail'

                # statu.append('Null')
            else:
                status.append('False')
                Diag = 'Parser fail'
                # statu.append('Null')

        if value.startswith("，"):
            Diag = 'KB is empty'
        if value == "是null":
            Diag = 'KB is empty'
        if value == "":
            Diag = 'KB is empty'
        if "有null" in value:
            Diag = 'KB is empty'
        di.append(Diag)

    file.close()

    if file_type == 1:
        save_file_name = file_name.split(".txt")
    else:
        save_file_name = file_name.split(".csv")

    save_file = save_dir + save_file_name[0] + ".xlsx"
    style = xlwt.XFStyle()
    al = xlwt.Alignment()
    al.horz = 0x02
    al.vert = 0x01
    style.alignment = al
    # 新建一个Excel文件
    wbook = openpyxl.Workbook()
    # 给Excel添加表头
    # 将JudgeValue去掉
    # Question为1，Diagnosis为2，Entity为3，ExpectEntity为4，JudgeEntity为5，Intent为6，ExpectIntent为7，JudgeIntent为8，Value为9
    biaotou = ['Question', 'Diagnosis', 'Entity', 'ExpectEntity', 'JudgeEntity', 'Intent', 'ExpectIntent',
               'JudgeIntent', 'Value', 'ParserEngineType', 'ParserTime', 'AnswerTime', 'ConceptType',
               'parameter & parameterType']
    bsheet = wbook.create_sheet('message', index=0)
    for i, item in enumerate(biaotou):
        bsheet.cell(row=1, column=i + 1, value=item)

    # 写入数据bsheet.write(行,列,value)
    bsheet.cell(2, 1, )
    bsheet.cell(2, 2, )
    bsheet.cell(2, 3, )
    bsheet.cell(2, 4, )
    bsheet.cell(2, 5, )
    bsheet.cell(2, 6, )
    bsheet.cell(2, 7, )
    bsheet.cell(2, 8, )
    bsheet.cell(2, 9, )
    bsheet.cell(2, 10, )
    bsheet.cell(2, 11, )
    bsheet.cell(2, 12)
    bsheet.cell(2, 13, )
    bsheet.cell(2, 14, )
    # bsheet.col(0).width = 256 * 28
    # bsheet.col(1).width = 256 * 35
    # bsheet.col(2).width = 256 * 20
    # bsheet.col(3).width = 256 * 33
    # bsheet.col(4).width = 256 * 60

    # 将问题  qu列表   写入到Excel中
    for i in range(len(qu)):
        bsheet.cell(i + 2, 1, qu[i])
    # 将 请求的状态 di列表 写入到Excel中
    for j in range(len(di)):
        bsheet.cell(j + 2, 2, di[j])

    # 将 实体列表 en列表 写入到Excel中
    for h in range(len(en)):
        bsheet.cell(h + 2, 3, en[h])

    # 将期望的实体 exp_ent列表 写入到Excel中
    if test_level == 1:
        for o in range(len(exp_ent)):
            bsheet.cell(o + 2, 4, exp_ent[o])

    # 将判断的 entity 写入到Excel中
    for q in range(len(en_statu)):
        bsheet.cell(q + 2, 5, en_statu[q])

    # 将Intent 写入到Excel中
    for p in range(len(inte)):
        bsheet.cell(p + 2, 6, inte[p])

    # 将 期望的意图 ExpectIntent 写入到Excel中
    for e in range(len(exp_int)):
        bsheet.cell(e + 2, 7, exp_int[e])

    # 将判断的意图 写入到Excel中
    for m in range(len(status)):
        bsheet.cell(m + 2, 8, status[m])

    # 将value 值写入到Excel中
    for l in range(len(va)):
        # if len(va) > 50:
        # bsheet.cell(l + 2, 9, va[l][:30])
        bsheet.cell(l + 2, 9, va[l])

    # 将ParserEngineType 值写入到Excel中
    for s in range(len(ParserEngineType)):
        bsheet.cell(s + 2, 10, ParserEngineType[s])

    # parser_time
    for a in range(len(parser_time)):
        bsheet.cell(a + 2, 11, parser_time[a])
    # answer_time
    for b in range(len(answer_time)):
        bsheet.cell(b + 2, 12, answer_time[b])
    # concept_type
    for j in range(len(concept_type)):
        bsheet.cell(j + 2, 13, concept_type[j])
    # parameter
    for x in range(len(parameter)):
        bsheet.cell(x + 2, 14, parameter[x])

    wbook.save(save_file)

    from openpyxl.styles import Alignment

    # 读取文件 设置列宽 表头居中
    wb = load_workbook(save_file)
    ws_list = wb.sheetnames

    for i in range(len(ws_list)):
        ws = wb[ws_list[i]]
        ws.column_dimensions["A"].width = 28.0
        ws.cell(1, 1, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["B"].width = 12.0
        ws.cell(1, 2, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["C"].width = 15.0
        ws.cell(1, 3, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["D"].width = 15.0
        ws.cell(1, 4, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["E"].width = 8.0
        ws.cell(1, 5, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["F"].width = 20.0
        ws.cell(1, 6, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["G"].width = 20.0
        ws.cell(1, 7, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["H"].width = 8.0
        ws.cell(1, 8, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["I"].width = 50.0
        ws.cell(1, 9, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["J"].width = 8.0
        ws.cell(1, 10, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["K"].width = 5.0
        ws.cell(1, 11, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["L"].width = 5.0
        ws.cell(1, 12, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["M"].width = 10.0
        ws.cell(1, 13, ).alignment = Alignment(horizontal="center")
        ws.column_dimensions["N"].width = 25.0
        ws.cell(1, 14, ).alignment = Alignment(horizontal="center")
        wb.save(save_file)


# get names of files in the directory
def get_filename(filepath, filetype):
    import os
    filename = []
    for root, dirs, files in os.walk(filepath):
        for i in files:
            if filetype in i:
                filename.append(i)
    return filename


if __name__ == '__main__':
    global file_dir
    file_dir = os.getcwd() + "/testcase/"
    if file_type == 1:
        filetype = '.txt'
    else:
        filetype = '.csv'

    # filenames = []
    # for root, dirs, files in os.walk(file_dir):
    #    for i in files:
    #        if filetype in i:
    #            filenames.append(i)

    # 新建的目录路径
    global save_dir
    save_dir = os.getcwd() + "/output/"
    if os.path.isdir(save_dir):
        print("output had been created")
    else:
        os.mkdir(save_dir)  # 新建目录

    # threads = []
    # process = []
    filenames = os.listdir(file_dir)

    pool = Pool(5)
    # for i in range(len(filenames)):
    pool.map(starttestcase, filenames)

    # p = Process(target=starttestcase,args=(filename[i], srcfilepath,desfilepath))
    # process.append(p)
    # t = threading.Thread(target=starttestcase, args=(filename[i], srcfilepath,desfilepath))
    # threads.append(t)
    # starttestcase(filename[i],srcfilepath,desfilepath)
    # for i in range(len(filename)):
    #    threads[i].start()

    # for i in range(len(filename)):
    #    threads[i].join()
    # pool.shutdown(wait=True)
